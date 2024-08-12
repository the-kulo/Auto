import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

# 读取本地文件
file_path_0 = r'D:\OneDrive\桌面\ebay\eBay IS_report_monthly report20240708_0000-20240715_0000.xlsx'

# 本地保存路径
file_path_1 = r'D:\OneDrive\桌面\ebay\ebay.xlsx'

# 加载工作簿
book = load_workbook(file_path_1)

# 删除不需要的工作表
for sheet in book.sheetnames:
    if sheet not in ['ECS', 'CPU Max Top 8', 'CPU Avg Top 8', 'Mem Max Top 8', 'Mem Avg Top 8']:
        del book[sheet]

# 创建工作表
if 'CPU Max Top 8' not in book.sheetnames:
    book.create_sheet(title='CPU Max Top 8')
if 'CPU Avg Top 8' not in book.sheetnames:
    book.create_sheet(title='CPU Avg Top 8')
if 'Mem Max Top 8' not in book.sheetnames:
    book.create_sheet(title='Mem Max Top 8')
if 'Mem Avg Top 8' not in book.sheetnames:
    book.create_sheet(title='Mem Avg Top 8')

# 保存工作簿
book.save(file_path_1)

# 读取数据
data1 = pd.read_excel(file_path_0, sheet_name='Linux Hosts')

# 筛选和列命名
data2 = pd.read_excel(file_path_0, usecols=[0, 2, 3, 4, 5])
data2.columns = ['主机', 'CPU Max Top 8', 'CPU Avg Top 8', 'Mem Max Top 8', 'Mem Avg Top 8']

# 使用 ExcelWriter 写入数据
with pd.ExcelWriter(file_path_1, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    # 写入筛选后的数据
    data2.to_excel(writer, sheet_name='ECS', index=False)

    # 筛选和写入 CPU Max Top 8
    CPU_MAX_top_8 = data2.sort_values(by=['CPU Max Top 8'], ascending=False).head(8)
    CPU_MAX_top_8_SELECTED = CPU_MAX_top_8[['主机', 'CPU Max Top 8']]
    CPU_MAX_top_8_SELECTED.to_excel(writer, sheet_name='CPU Max Top 8', index=False)

    # 筛选和写入 CPU Avg Top 8
    CPU_AVG_top_8 = data2.sort_values(by=['CPU Avg Top 8'], ascending=False).head(8)
    CPU_AVG_top_8_SELECTED = CPU_AVG_top_8[['主机', 'CPU Avg Top 8']]
    CPU_AVG_top_8_SELECTED.to_excel(writer, sheet_name='CPU Avg Top 8', index=False)

    # 筛选和写入 Mem Max Top 8
    MEM_MAX_top_8 = data2.sort_values(by=['Mem Max Top 8'], ascending=False).head(8)
    MEM_MAX_top_8_SELECTED = MEM_MAX_top_8[['主机', 'Mem Max Top 8']]
    MEM_MAX_top_8_SELECTED.to_excel(writer, sheet_name='Mem Max Top 8', index=False)

    # 筛选和写入 Mem Avg Top 8
    MEM_AVG_top_8 = data2.sort_values(by=['Mem Avg Top 8'], ascending=False).head(8)
    MEM_AVG_top_8_SELECTED = MEM_AVG_top_8[['主机', 'Mem Avg Top 8']]
    MEM_AVG_top_8_SELECTED.to_excel(writer, sheet_name='Mem Avg Top 8', index=False)

# 重新加载工作簿
book = load_workbook(file_path_1)

# 添加图表
def add_chart(sheet_name, value_column, title):
    sheet = book[sheet_name]
    chart = BarChart()
    chart.title = title
    data = Reference(sheet, min_col=value_column, min_row=1, max_col=value_column, max_row=9)
    categories = Reference(sheet, min_col=1, min_row=2, max_row=9)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    sheet.add_chart(chart, "E5")

# 添加图表到对应工作表
add_chart('CPU Max Top 8', 2, 'CPU Max Top 8')
add_chart('CPU Avg Top 8', 2, 'CPU Avg Top 8')
add_chart('Mem Max Top 8', 2, 'Mem Max Top 8')
add_chart('Mem Avg Top 8', 2, 'Mem Avg Top 8')

# 保存工作簿
book.save(file_path_1)
