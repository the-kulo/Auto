import pandas as pd
from openpyxl import load_workbook

# 读取Excel文件
file_path = 'D:\\OneDrive\\桌面\\Exchange\\Monthly.xlsx'
df = pd.read_excel(file_path)

# 筛选出Severity列中值为4或5的行
filtered_df = df[df['Severity'].isin([4, 5])]

# 将筛选结果写入新的Excel文件
output_file_path = 'D:\\OneDrive\\桌面\\Exchange\\Monthly-scan.xlsx'
filtered_df.to_excel(output_file_path, index=False)

# 使用openpyxl加载工作簿
wb = load_workbook(output_file_path)

# 创建Sheet2和Sheet3工作表
sheet2 = wb.create_sheet(title='UAT')
sheet3 = wb.create_sheet(title='Prod')

# 设置需要保留的列
columns = ['QID', 'Title', 'Severity', 'IP', 'DNS', 'OS', 'Category',
           'Vendor Reference', 'Impact', 'Solution', 'Results']

# 重命名列的字典
column_mapping = {
    'QID': 'QID',
    'Title': '漏洞标题',
    'Severity': '漏洞级别',
    'IP': '影响资产IP',
    'DNS': '影响资产备注名称',
    'OS': 'OS',
    'Category': '类别',
    'Vendor Reference': 'Vendor Reference',
    'Impact': '受影响',
    'Solution': '建议的解决方案',
    'Results': 'Results'
}

# 写入列标题到Sheet2和Sheet3的第一行
for col_num, column in enumerate(columns, 1):
    sheet2.cell(row=1, column=col_num, value=column_mapping[column])
    sheet3.cell(row=1, column=col_num, value=column_mapping[column])

# 添加三列分别为环境，修复途径，备注
extra_columns = ['环境', '修复途径', '备注']
for col_num, column in enumerate(extra_columns, len(columns) + 1):
    sheet2.cell(row=1, column=col_num, value=column)
    sheet3.cell(row=1, column=col_num, value=column)

# 筛选DNS列中包含'-d'和'-p'的行
dns_d_df = filtered_df[filtered_df['DNS'].str.contains('-d', na=False)].copy()
dns_p_df = filtered_df[filtered_df['DNS'].str.contains('-p', na=False)].copy()

# 提取环境列数据
def extract_environment(dns_value):
    if '-' in dns_value and '.' in dns_value:
        start = dns_value.find('-') + 1
        end = dns_value.find('.', start)
        return dns_value[start:end]
    return ""

dns_d_df['环境'] = dns_d_df['DNS'].apply(extract_environment)
dns_p_df['环境'] = dns_p_df['DNS'].apply(extract_environment)

# 根据Results列中的关键词提取修复途径和备注
def detect_keywords(result_value):
    # 将结果值转换为小写
    result_value = result_value.lower()
    if 'log4j' in result_value:
        return 'log4j升级', '升级会导致hybris无法运行，漏洞以避免，不建议升级'
    elif 'package' in result_value:
        return '升级rpm包', ''
    elif 'jdk' in result_value:
        return 'openjdk升级', 'jdk版本升级影响面比较大，改动也会比较大，不建议升级'
    elif 'php' in result_value:
        return 'PHP', '升级会导致PHP无法运行，漏洞以避免，不建议升级'
    elif 'apache' in result_value:
        return 'Apache', '需要应用查看Apache版本进行排查更新'
    return '', ''

# 添加修复途径和备注列
dns_d_df[['修复途径', '备注']] = dns_d_df['Results'].apply(lambda x: pd.Series(detect_keywords(x)))
dns_p_df[['修复途径', '备注']] = dns_p_df['Results'].apply(lambda x: pd.Series(detect_keywords(x)))

# 将筛选后的数据写入Sheet2和Sheet3
for row_num, row_data in enumerate(dns_d_df[columns + ['环境', '修复途径', '备注']].values, 2):
    for col_num, cell_value in enumerate(row_data, 1):
        sheet2.cell(row=row_num, column=col_num, value=cell_value)

for row_num, row_data in enumerate(dns_p_df[columns + ['环境', '修复途径', '备注']].values, 2):
    for col_num, cell_value in enumerate(row_data, 1):
        sheet3.cell(row=row_num, column=col_num, value=cell_value)

# 保存工作簿
wb.save(output_file_path)

print("筛选后的数据已成功保存到新的Excel文件中。")
